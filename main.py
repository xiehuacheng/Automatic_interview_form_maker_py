import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
from tkinter import messagebox
import openpyxl


class Person:
    def __init__(self, name, preferred_slot, availability, personal_info):
        self.name = name
        self.preferred_slot = preferred_slot
        self.availability = availability
        self.personal_info = personal_info
        self.assigned_slot = None


def read_excel(file_path, num_slots):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    students = []
    personal_info_header = []
    for row in sheet.iter_rows(min_row=1, max_row=1):
        personal_info_header = [cell.value for cell in row[num_slots+2:]]

    for row in sheet.iter_rows(min_row=2):
        name = row[0].value
        preferred_slot = row[1].value
        availability = [cell.value for cell in row[2:num_slots+2]]
        personal_info = [cell.value for cell in row[num_slots+2:]]
        students.append(Person(name, preferred_slot,
                        availability, personal_info))

    return students, personal_info_header


def write_excel(data, file_path, personal_info_header):
    wb = openpyxl.Workbook()
    sheet = wb.active

    # 写入表头
    slots = len(data[0].availability)
    header = ['姓名', '安排时间段', '首选时间段'] + \
        ['时间段{}'.format(i + 1) for i in range(slots)] + personal_info_header
    sheet.append(header)

    # 写入学生信息
    for student in data:
        row_data = [student.name, student.assigned_slot,
                    student.preferred_slot] + student.availability + student.personal_info
        sheet.append(row_data)

    wb.save(file_path)


def schedule_interviews(students, max_students_per_slot):
    assigned_students = []
    unassigned_students = []
    schedule = [[] for _ in range(len(max_students_per_slot))]

    # 按首选时间段对学生进行排序
    students.sort(
        key=lambda s: s.preferred_slot if s.preferred_slot is not None else -1, reverse=True)

    # 将学生分配到时间段
    for student in students:
        assigned = False

        # 检查学生是否只有一个可用时间段
        if sum(student.availability) == 1:
            slot = student.availability.index(1) + 1
            if len(schedule[slot-1]) < max_students_per_slot[slot-1]:
                student.assigned_slot = slot
                schedule[slot-1].append(student)
                assigned_students.append(student)
                assigned = True

        # 检查学生是否有多个可用时间段
        else:
            preferred_slot = student.preferred_slot
            if preferred_slot is not None and student.availability[preferred_slot-1] and len(schedule[preferred_slot-1]) < max_students_per_slot[preferred_slot-1]:
                student.assigned_slot = preferred_slot
                schedule[preferred_slot-1].append(student)
                assigned_students.append(student)
                assigned = True
            else:
                available_slots = [slot for slot, is_available in enumerate(
                    student.availability, start=1) if is_available and len(schedule[slot-1]) < max_students_per_slot[slot-1]]
                if available_slots:
                    student.assigned_slot = available_slots[0]
                    schedule[available_slots[0]-1].append(student)
                    assigned_students.append(student)
                    assigned = True
                else:
                    # 替换同一时间段的学生
                    print(
                        f"Trying to replace a student in the same slot for {student.name}")
                    possible_slots = [slot for slot, is_available in enumerate(
                        student.availability, start=1) if is_available]
                    for slot in possible_slots:
                        for assigned_student in schedule[slot-1]:
                            otherAvailable_slots = [tempSlot for tempSlot, is_available in enumerate(
                                assigned_student.availability, start=1) if is_available and len(schedule[tempSlot-1]) < max_students_per_slot[tempSlot-1]]
                            if otherAvailable_slots:
                                print(
                                    f"Replaced {assigned_student.name} with {student.name}")
                                assigned_student.assigned_slot = otherAvailable_slots[0]
                                student.assigned_slot = slot
                                schedule[slot-1].append(student)
                                schedule[slot-1].remove(assigned_student)
                                schedule[otherAvailable_slots[0] -
                                         1].append(assigned_student)
                                assigned_students.append(student)
                                assigned = True
                                break

                        if assigned:
                            break

                    if not assigned:
                        unassigned_students.append(student)

    return assigned_students, unassigned_students


def browse_file():
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel Files", "*.xlsx")])
    entry_file.delete(0, tk.END)
    entry_file.insert(tk.END, file_path)


def generate_schedule():
    file_path = entry_file.get()
    num_slots = int(slot_combobox.get())

    max_students_per_slot = []
    for i in range(num_slots):
        entry = max_students_entries[i]
        max_students_per_slot.append(int(entry[1].get()))

    students, personal_info_header = read_excel(file_path, num_slots)

    assigned_students, unassigned_students = schedule_interviews(
        students, max_students_per_slot)

    schedule = [[] for _ in range(num_slots)]
    for student in assigned_students:
        schedule[student.assigned_slot - 1].append(student)

    if len(unassigned_students) == 0:
        write_excel(assigned_students, "schedule.xlsx", personal_info_header)
        lbl_result.config(text="成功生成面试时间安排！")
    else:
        message = "以下学生未能被安排到合适的时间段：\n"
        for student in unassigned_students:
            message += f"{student.name}\n"
        message += "请增加以下时间段的最大学生数量：\n"
        for i, slot in enumerate(max_students_per_slot):
            if len(schedule[i]) == slot:
                message += f"时间段 {i+1}：还需增加 {sum(student.availability[i] for student in unassigned_students)} 名学生\n"
        messagebox.showwarning("警告", message)


# 创建图形化界面
window = tk.Tk()
window.title("自动面试时间生成器")

lbl_file = tk.Label(window, text="Excel 文件：")
lbl_file.pack()
entry_file = tk.Entry(window, width=50)
entry_file.pack()
btn_browse = tk.Button(window, text="选择文件", command=browse_file)
btn_browse.pack()

lbl_slots = tk.Label(window, text="面试时间段数：")
lbl_slots.pack()
slot_combobox = ttk.Combobox(
    window, values=[str(i) for i in range(2, 8)])  # 可选择2~7个时间段
slot_combobox.pack()

lbl_max_students = tk.Label(window, text="每个时间段最大学生数：")
lbl_max_students.pack()

max_students_entries = []


def create_max_students_entries():
    num_slots = int(slot_combobox.get())

    # 清除前一次选择的文本
    for entry in max_students_entries:
        entry[0].destroy()
        entry[1].destroy()
    max_students_entries.clear()

    # 创建新的输入框和文本
    for i in range(num_slots):
        lbl_max_students = tk.Label(window, text=f"时间段 {i + 1} 最大学生数：")
        lbl_max_students.pack()
        entry_max_students = tk.Entry(window, width=10)
        entry_max_students.pack()
        max_students_entries.append((lbl_max_students, entry_max_students))


btn_generate = tk.Button(window, text="生成面试时间安排", command=generate_schedule)
btn_generate.pack()

lbl_result = tk.Label(window, text="")
lbl_result.pack()

slot_combobox.bind("<<ComboboxSelected>>",
                   lambda event: create_max_students_entries())

window.mainloop()
